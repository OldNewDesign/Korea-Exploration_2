"""Excel export from SQLite. One overview tab, a tab per category present,
plus Needs Review and Errors tabs."""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from . import config, store

# openpyxl rejects most ASCII control chars (e.g. ANSI color codes in error text)
_ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _clean(v):
    return _ILLEGAL.sub("", v) if isinstance(v, str) else v

# (db_column, header, width)
DISPLAY = [
    ("platform", "Platform", 12), ("creator", "Creator", 22),
    ("topic", "Category", 20), ("sub_category", "Subcategory", 18),
    ("location", "Location", 34), ("summary", "Summary", 50),
    ("caption_english", "Caption (EN)", 44), ("caption_original", "Caption (orig)", 36),
    ("detected_language", "Lang", 10), ("hashtags", "Hashtags", 28),
    ("keywords", "Keywords", 26), ("usefulness", "Use", 6),
    ("action", "Action", 14), ("confidence", "Conf", 7),
    ("view_count", "Views", 11), ("like_count", "Likes", 10),
    ("comment_count", "Comments", 10), ("duration", "Secs", 7),
    ("upload_date", "Uploaded", 11), ("url", "URL", 40),
]

HDR_FILL = PatternFill("solid", fgColor="C8472F")
HDR_FONT = Font(bold=True, color="FDFAF3", name="Calibri", size=11)
REVIEW_FILL = PatternFill("solid", fgColor="FBE6DF")
THIN = Side(style="thin", color="E2D8C6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")


def _safe_sheet_name(name, used):
    s = re.sub(r"[\[\]\:\*\?\/\\]", " ", name).strip()[:31] or "Sheet"
    base, i = s, 2
    while s.lower() in used:
        suffix = f" ({i})"
        s = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(s.lower())
    return s


def _write_table(ws, rows):
    for c, (_, header, width) in enumerate(DISPLAY, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.fill, cell.font, cell.border = HDR_FILL, HDR_FONT, BORDER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = width
    for r, row in enumerate(rows, start=2):
        review = row.get("needs_review")
        for c, (col, _, _) in enumerate(DISPLAY, start=1):
            cell = ws.cell(row=r, column=c, value=_clean(row.get(col, "")))
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if review:
                cell.fill = REVIEW_FILL
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(DISPLAY))}{len(rows)+1}"


def build(path=None):
    path = str(path or config.EXCEL_PATH)
    rows = store.all_videos()
    errors = store.all_errors()

    wb = Workbook()
    used = set()

    ws_all = wb.active
    ws_all.title = _safe_sheet_name("All Videos", used)
    _write_table(ws_all, rows)

    # one tab per category, in config order, only those present
    by_cat = {}
    for row in rows:
        by_cat.setdefault(row.get("topic") or "Other", []).append(row)
    order = [c for c in config.CATEGORIES if c in by_cat]
    order += [c for c in by_cat if c not in order]
    for cat in order:
        ws = wb.create_sheet(_safe_sheet_name(cat, used))
        _write_table(ws, by_cat[cat])

    ws_rev = wb.create_sheet(_safe_sheet_name("Needs Review", used))
    _write_table(ws_rev, [r for r in rows if r.get("needs_review")])

    ws_err = wb.create_sheet(_safe_sheet_name("Errors", used))
    headers = ["URL", "Stage", "Message", "When"]
    widths = [42, 14, 70, 20]
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws_err.cell(row=1, column=c, value=h)
        cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CENTER, BORDER
        ws_err.column_dimensions[get_column_letter(c)].width = w
    for r, e in enumerate(errors, start=2):
        for c, key in enumerate(["url", "stage", "message", "ts"], start=1):
            cell = ws_err.cell(row=r, column=c, value=_clean(e.get(key, "")))
            cell.alignment, cell.border = WRAP_TOP, BORDER
    ws_err.freeze_panes = "A2"

    wb.save(path)
    return path
